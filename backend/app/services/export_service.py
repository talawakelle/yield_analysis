from __future__ import annotations

from pathlib import Path
import zipfile

import pandas as pd

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, A3
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


REGION_ALIASES = {
    "LC": "TTEL_LC",
    "RB": "HT",
}


def _canonical_region(region: str) -> str:
    return REGION_ALIASES.get(str(region).strip().upper(), str(region).strip().upper())


def _company_for_region(region: str) -> str:
    region = _canonical_region(region)
    if region in {"TK", "TTEL_LC", "NO"}:
        return "TTEL"
    if region in {"NE", "HT", "KVPL_LC"}:
        return "KVPL"
    if region in {"UC", "LD", "HPL_LC"}:
        return "HPL"
    return "TTEL"


COMPANY_EXCEL_STYLE = {
    "TTEL": {
        "header_fill": "E6E6E6",
        "header_font_color": "000000",
        "benchmark_fill": "F3D4B3",
        "bm90_fill": "CDEECB",
        "body_font": "Times New Roman",
    },
    "KVPL": {
        "header_fill": "E3E3E3",
        "header_font_color": "000000",
        "benchmark_fill": "F3D4B3",
        "bm90_fill": "CDEECB",
        "body_font": "Times New Roman",
    },
    "HPL": {
        "header_fill": "E3E3E3",
        "header_font_color": "000000",
        "benchmark_fill": "F3D4B3",
        "bm90_fill": "CDEECB",
        "body_font": "Times New Roman",
    },
}


def ensure_dir(path: str) -> Path:
    p = Path(path)
    p.mkdir(parents=True, exist_ok=True)
    return p


def _safe_sheet_name(region: str, sheet_name: str) -> str:
    return f"{region}_{sheet_name}"[:31]


def _style_worksheet(ws, df: pd.DataFrame, region: str) -> None:
    company = _company_for_region(region)
    style_cfg = COMPANY_EXCEL_STYLE[company]

    header_fill = PatternFill(fill_type="solid", fgColor=style_cfg["header_fill"])
    benchmark_fill = PatternFill(fill_type="solid", fgColor=style_cfg["benchmark_fill"])
    bm90_fill = PatternFill(fill_type="solid", fgColor=style_cfg["bm90_fill"])

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font = Font(
        bold=True,
        name=style_cfg["body_font"],
        size=11,
        color=style_cfg["header_font_color"],
    )
    body_font = Font(name=style_cfg["body_font"], size=10)

    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    max_row = ws.max_row
    max_col = ws.max_column

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center

    for row in ws.iter_rows(min_row=2, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font = body_font
            cell.border = border
            cell.alignment = center

    for col_cells in ws.columns:
        first_cell = col_cells[0]
        col_letter = get_column_letter(first_cell.column)

        max_len = 0
        for cell in col_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))

        if col_letter == "A":
            width = min(max(max_len + 2, 12), 24)
        else:
            width = min(max(max_len + 2, 10), 28)
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 24
    for r in range(2, max_row + 1):
        ws.row_dimensions[r].height = 20

    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False
    ws.auto_filter.ref = ws.dimensions
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_title_rows = "1:1"
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.40
    ws.page_margins.bottom = 0.40
    ws.page_margins.header = 0.20
    ws.page_margins.footer = 0.20

    col_names = [str(c).strip() if c is not None else "" for c in df.columns.tolist()]
    col_index = {name: idx + 1 for idx, name in enumerate(col_names)}

    benchmark_col = col_index.get("Benchmark")
    bm90_col = col_index.get("90% BM")

    estate_yph_cols = []
    for col in col_names:
        if col in {"Year", "Ext", "% Ext", "Benchmark", "90% BM"}:
            continue
        if str(col).endswith("_PCT") or str(col).endswith(" % EXT"):
            continue
        estate_yph_cols.append(col)

    if benchmark_col is not None:
        for excel_row in range(2, max_row + 1):
            bm_cell = ws.cell(row=excel_row, column=benchmark_col)
            try:
                benchmark_val = float(str(bm_cell.value).replace(",", ""))
            except Exception:
                benchmark_val = None

            try:
                bm90_val = float(str(ws.cell(row=excel_row, column=bm90_col).value).replace(",", "")) if bm90_col else None
            except Exception:
                bm90_val = None

            if benchmark_val is not None:
                bm_cell.fill = benchmark_fill

            for col_name in estate_yph_cols:
                col_num = col_index.get(col_name)
                if col_num is None:
                    continue

                cell = ws.cell(row=excel_row, column=col_num)
                try:
                    val = float(str(cell.value).replace(",", ""))
                except Exception:
                    continue

                if benchmark_val is not None and val == benchmark_val:
                    cell.fill = benchmark_fill
                elif bm90_val is not None and val >= bm90_val:
                    cell.fill = bm90_fill


def export_excel(result_payload: dict, output_dir: str) -> str:
    out_dir = ensure_dir(output_dir)
    path = out_dir / "report.xlsx"

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for region, sheets in result_payload.items():
            region_key = _canonical_region(region)

            for sheet_name, obj in sheets.items():
                safe_sheet = _safe_sheet_name(region_key, sheet_name)
                df = obj.copy() if isinstance(obj, pd.DataFrame) else pd.DataFrame(obj)
                df.to_excel(writer, sheet_name=safe_sheet, index=False)

                ws = writer.book[safe_sheet]
                _style_worksheet(ws, df, region_key)

    return str(path)



def export_pdf_from_images(image_paths: list[str], output_dir: str, filename: str = "report.pdf") -> str:
    out_dir = ensure_dir(output_dir)
    path = out_dir / filename

    c = canvas.Canvas(str(path), pagesize=landscape(A3))
    page_w, page_h = landscape(A3)

    margin = 4 * mm
    max_w = page_w - (2 * margin)
    max_h = page_h - (2 * margin)

    for img in image_paths:
        reader = ImageReader(img)
        img_w, img_h = reader.getSize()

        scale = min(max_w / img_w, max_h / img_h)
        draw_w = img_w * scale
        draw_h = img_h * scale

        x = (page_w - draw_w) / 2
        y = (page_h - draw_h) / 2

        c.setFillColor(colors.white)
        c.rect(0, 0, page_w, page_h, stroke=0, fill=1)
        c.drawImage(
            reader,
            x,
            y,
            width=draw_w,
            height=draw_h,
            preserveAspectRatio=True,
            anchor="c",
            mask="auto",
        )
        c.showPage()

    c.save()
    return str(path)


def export_abstract_table_pdf(df: pd.DataFrame, output_dir: str, filename: str = "abstract_summary.pdf") -> str:
    out_dir = ensure_dir(output_dir)
    path = out_dir / filename

    doc = SimpleDocTemplate(
        str(path),
        pagesize=landscape(A3),
        leftMargin=8,
        rightMargin=8,
        topMargin=10,
        bottomMargin=10,
    )

    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("<b>Abstract Summary</b>", styles["Title"]))
    elements.append(Spacer(1, 6))

    table_data, spans = _build_pdf_table_from_dataframe(df)

    page_width, _ = landscape(A3)
    usable_width = page_width - (16 * mm)
    col_count = len(table_data[0])

    if col_count >= 3:
        fixed_widths = [30 * mm, 27 * mm, 24 * mm]
        remaining_count = max(col_count - 3, 1)
        remaining_width = max((usable_width - sum(fixed_widths)) / remaining_count, 16 * mm)
        col_widths = fixed_widths + [remaining_width] * (col_count - 3)
    else:
        col_widths = [usable_width / max(col_count, 1)] * col_count

    table = Table(table_data, colWidths=col_widths, repeatRows=2)

    style_cmds = [
        ("GRID", (0, 0), (-1, -1), 0.8, colors.black),
        ("BOX", (0, 0), (-1, -1), 1.2, colors.black),

        ("BACKGROUND", (0, 0), (-1, 1), colors.HexColor("#E6E6E6")),
        ("FONTNAME", (0, 0), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 1), 8),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),

        ("FONTNAME", (0, 2), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 2), (-1, -1), 7.5),

        ("TOPPADDING", (0, 0), (-1, 1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, 1), 7),
        ("TOPPADDING", (0, 2), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 2), (-1, -1), 4),

        ("FONTNAME", (0, 2), (0, -1), "Helvetica-Bold"),
    ]

    for span in spans:
        style_cmds.append(("SPAN", span[0], span[1]))

    _apply_highlight_styles(table_data, style_cmds)

    table.setStyle(TableStyle(style_cmds))
    elements.append(table)
    elements.append(Spacer(1, 6))
    elements.append(
        Paragraph(
            "<b>Peach - Benchmark | Green - YPH Greater than 90% of Benchmark</b>",
            styles["Normal"],
        )
    )

    doc.build(elements)
    return str(path)


def _build_pdf_table_from_dataframe(df: pd.DataFrame):
    cols = [str(c).strip() for c in df.columns.tolist()]

    row1 = []
    row2 = []
    spans = []

    fixed = cols[:3]
    for col in fixed:
        start = len(row1)
        row1.append(col)
        row2.append("")
        spans.append(((start, 0), (start, 1)))

    middle_cols = cols[3:-2]
    i = 0
    while i < len(middle_cols):
        left_col = middle_cols[i]

        estate_name = (
            left_col.replace(" YPH", "")
            .replace("YPH", "")
            .replace("% EXT", "")
            .replace("%EXT", "")
            .replace("_PCT", "")
            .replace("_YPH", "")
            .strip()
        )

        start = len(row1)
        row1.extend([estate_name, estate_name])
        row2.extend(["YPH", "% EXT"])
        spans.append(((start, 0), (start + 1, 0)))

        i += 2

    trailing = cols[-2:]
    for col in trailing:
        start = len(row1)
        row1.append(col)
        row2.append("")
        spans.append(((start, 0), (start, 1)))

    table_data = [row1, row2]

    for _, row in df.iterrows():
        table_data.append([_safe_cell(v) for v in row.tolist()])

    return table_data, spans


def _safe_cell(value):
    if pd.isna(value):
        return ""
    return str(value)


def _apply_highlight_styles(table_data, style_cmds):
    if len(table_data) < 3 or len(table_data[0]) < 6:
        return

    header_row_1 = table_data[0]
    header_row_2 = table_data[1]

    yph_cols = []
    benchmark_col = None
    bm90_col = None

    for idx, (h1, h2) in enumerate(zip(header_row_1, header_row_2)):
        if str(h2).strip().upper() == "YPH":
            yph_cols.append(idx)
        if str(h1).strip().lower() == "benchmark":
            benchmark_col = idx
        if str(h1).strip().lower() == "90% bm":
            bm90_col = idx

    if benchmark_col is None or not yph_cols:
        return

    for r in range(2, len(table_data)):
        benchmark_val = None
        bm90_val = None

        try:
            benchmark_val = float(str(table_data[r][benchmark_col]).replace(",", ""))
        except Exception:
            benchmark_val = None

        try:
            if bm90_col is not None:
                bm90_val = float(str(table_data[r][bm90_col]).replace(",", ""))
        except Exception:
            bm90_val = None

        if benchmark_val is not None:
            style_cmds.append(
                ("BACKGROUND", (benchmark_col, r), (benchmark_col, r), colors.HexColor("#F3D4B3"))
            )

        for c in yph_cols:
            try:
                val = float(str(table_data[r][c]).replace(",", ""))
            except Exception:
                continue

            if benchmark_val is not None and val == benchmark_val:
                style_cmds.append(
                    ("BACKGROUND", (c, r), (c, r), colors.HexColor("#F3D4B3"))
                )
            elif bm90_val is not None and val >= bm90_val:
                style_cmds.append(
                    ("BACKGROUND", (c, r), (c, r), colors.HexColor("#CDEECB"))
                )


def export_zip_package(output_dir: str, zip_name: str = "report_package.zip") -> str:
    out_dir = ensure_dir(output_dir)
    zip_path = out_dir / zip_name

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for p in out_dir.rglob("*"):
            if p.is_file() and p.name != zip_name:
                zf.write(p, arcname=str(p.relative_to(out_dir)))

    return str(zip_path)